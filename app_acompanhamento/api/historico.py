import json
import os
from io import BytesIO

import psycopg2.extras
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from src.db import Historico

router = APIRouter()


def _dsn_ou_erro() -> str:
    dsn = os.environ.get("POSTGRES_DSN")
    if not dsn:
        raise HTTPException(400, "POSTGRES_DSN nao configurado")
    return dsn


@router.get("/historico")
def historico():
    """So leitura - historico do envio diario (origem odoo_manual/odoo_automatico), sempre
    filtrado - essa tela nunca mostra o historico da importacao em massa via Excel."""
    dsn = _dsn_ou_erro()
    hist = Historico(dsn)
    try:
        with hist.conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM import_execucao WHERE origem LIKE %s ORDER BY id DESC LIMIT 20",
                ("odoo%",),
            )
            colunas = [d[0] for d in cur.description]
            rows = [dict(zip(colunas, row)) for row in cur.fetchall()]
    finally:
        hist.close()

    for row in rows:
        for key, value in row.items():
            if hasattr(value, "isoformat"):
                row[key] = value.isoformat()

    return {"execucoes": rows}


@router.get("/historico/{execucao_id}/linhas")
def historico_linhas(execucao_id: int, apenas_erros: bool = True):
    dsn = _dsn_ou_erro()
    hist = Historico(dsn)
    try:
        with hist.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            filtro_status = "AND status LIKE 'erro_%%'" if apenas_erros else ""
            cur.execute(
                f"""
                SELECT linha_excel, colaborador, data, projeto, idsap, tarefa, horas,
                       employee_id_sap, status, motivo, sap_response, sap_request
                FROM import_linha
                WHERE execucao_id = %s {filtro_status}
                ORDER BY linha_excel
                """,
                (execucao_id,),
            )
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        hist.close()

    for row in rows:
        for key, value in row.items():
            if hasattr(value, "isoformat"):
                row[key] = value.isoformat()

    return {"linhas": rows}


@router.get("/historico/exportar")
def exportar_historico_completo(data_inicio: str | None = None, data_fim: str | None = None):
    """So leitura - Excel consolidado de todas as execucoes do envio diario (origem odoo_*),
    com filtro opcional de periodo (por criado_em, quando cada linha foi processada)."""
    dsn = _dsn_ou_erro()
    hist = Historico(dsn)
    try:
        with hist.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            sql = """
                SELECT e.id AS execucao_id, e.origem, e.ambiente, e.iniciado_em, e.finalizado_em,
                       l.linha_excel, l.colaborador, l.data, l.projeto, l.idsap, l.tarefa,
                       l.descricao, l.horas, l.employee_id_sap, l.status, l.motivo, l.criado_em
                FROM import_linha l
                JOIN import_execucao e ON e.id = l.execucao_id
                WHERE e.origem LIKE %s
            """
            params: list = ["odoo%"]
            if data_inicio:
                sql += " AND l.criado_em::date >= %s"
                params.append(data_inicio)
            if data_fim:
                sql += " AND l.criado_em::date <= %s"
                params.append(data_fim)
            sql += " ORDER BY e.id, l.linha_excel"
            cur.execute(sql, params)
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        hist.close()

    if not rows:
        raise HTTPException(404, "Nenhuma linha encontrada no historico para o filtro informado")

    colunas = ["execucao_id", "origem", "ambiente", "iniciado_em", "finalizado_em", "linha_excel",
               "colaborador", "data", "projeto", "idsap", "tarefa", "descricao", "horas",
               "employee_id_sap", "status", "motivo", "criado_em"]

    wb = Workbook()
    ws = wb.active
    ws.title = "historico_diario"
    ws.append(colunas)
    for row in rows:
        linha = []
        for coluna in colunas:
            valor = row.get(coluna)
            if hasattr(valor, "isoformat"):
                valor = valor.isoformat()
            linha.append(valor)
        ws.append(linha)

    for i, coluna in enumerate(colunas, start=1):
        largura = 14 if coluna in ("status", "data", "idsap", "horas", "ambiente", "origem") else 22
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    sufixo = f"_{data_inicio or 'inicio'}_a_{data_fim or 'hoje'}" if (data_inicio or data_fim) else ""
    nome_arquivo = f"historico_diario{sufixo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/historico/{execucao_id}/exportar")
def exportar_execucao(execucao_id: int):
    """So leitura - Excel com todas as linhas de uma execucao especifica do envio diario."""
    dsn = _dsn_ou_erro()
    hist = Historico(dsn)
    try:
        with hist.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT l.linha_excel, l.colaborador, l.data, l.projeto, l.idsap, l.tarefa,
                       l.descricao, l.horas, l.employee_id_sap, l.status, l.motivo,
                       l.sap_request, l.sap_response, l.criado_em
                FROM import_linha l
                JOIN import_execucao e ON e.id = l.execucao_id
                WHERE l.execucao_id = %s AND e.origem LIKE 'odoo%%'
                ORDER BY l.linha_excel
                """,
                (execucao_id,),
            )
            rows = [dict(r) for r in cur.fetchall()]
    finally:
        hist.close()

    if not rows:
        raise HTTPException(404, f"Nenhuma linha encontrada para a execucao {execucao_id}")

    colunas = ["linha_excel", "colaborador", "data", "projeto", "idsap", "tarefa", "descricao",
               "horas", "employee_id_sap", "status", "motivo", "sap_request", "sap_response", "criado_em"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"execucao_{execucao_id}"
    ws.append(colunas)
    for row in rows:
        linha = []
        for coluna in colunas:
            valor = row.get(coluna)
            if isinstance(valor, (dict, list)):
                valor = json.dumps(valor, ensure_ascii=False)
            elif hasattr(valor, "isoformat"):
                valor = valor.isoformat()
            linha.append(valor)
        ws.append(linha)

    for i, coluna in enumerate(colunas, start=1):
        largura = 14 if coluna in ("status", "data", "idsap", "horas") else 24
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = largura

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"execucao_{execucao_id}_status.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
