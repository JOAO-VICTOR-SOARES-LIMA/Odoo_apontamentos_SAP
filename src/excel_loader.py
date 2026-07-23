import datetime
import openpyxl

# Cada coluna logica aceita mais de um nome de header - o export do n8n mudou de
# projeto/descricao/horas para projeto_odoo/descricao_apontamento/horas_apontadas
# em versoes mais novas do fluxo, mantendo o mesmo significado.
COLUNAS_ACEITAS = {
    "colaborador":     ["colaborador"],
    "data":            ["data"],
    "projeto":         ["projeto", "projeto_odoo"],
    "idsap":           ["idsap"],
    "tarefa":          ["tarefa"],
    "descricao":       ["descricao", "descricao_apontamento"],
    "horas":           ["horas", "horas_apontadas"],
    "StartTime":       ["StartTime"],
    "EndTime":         ["EndTime"],
    "Break":           ["Break"],
    "employeeId_sap":  ["employeeId_sap"],
}

# Coluna opcional: so vai pro historico/auditoria, nao afeta validacao - ausente em alguns exports.
COLUNA_INTERVALO_ACEITAS = ["intervalo"]

# Coluna opcional: se o Excel trouxer o codigo de projeto autoritativo do SAP
# (ex: campo x_studio_related_field do Odoo), usamos ele em vez de derivar de "projeto".
COLUNAS_PROJETO_SAP_ACEITAS = ["projeto_sap", "codigo_sap_projeto", "financial_project", "FinancialProject"]

# Colunas opcionais de reconciliacao - vem vazias no export atual, so exibicao/auditoria no front.
COLUNAS_SAP_RECONCILIACAO = {
    "sap_data":          ["sap_data"],
    "sap_hora_inicio":   ["sap_horaInicio"],
    "sap_hora_fim":      ["sap_horaFim"],
    "sap_observacao":    ["sap_observacao"],
}


def carregar_linhas(caminho_excel: str) -> list[dict]:
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value else None for c in ws[1]]

    idx = {}
    faltando = []
    for nome_logico, aceitos in COLUNAS_ACEITAS.items():
        coluna_encontrada = next((c for c in aceitos if c in header), None)
        if coluna_encontrada is None:
            faltando.append("/".join(aceitos))
        else:
            idx[nome_logico] = header.index(coluna_encontrada)
    if faltando:
        raise ValueError(f"Excel nao tem as colunas esperadas: {faltando}. Header encontrado: {header}")

    coluna_intervalo = next((c for c in COLUNA_INTERVALO_ACEITAS if c in header), None)
    idx_intervalo = header.index(coluna_intervalo) if coluna_intervalo else None

    coluna_projeto_sap = next((c for c in COLUNAS_PROJETO_SAP_ACEITAS if c in header), None)
    idx_projeto_sap = header.index(coluna_projeto_sap) if coluna_projeto_sap else None

    idx_reconciliacao = {}
    for nome_logico, aceitos in COLUNAS_SAP_RECONCILIACAO.items():
        coluna_encontrada = next((c for c in aceitos if c in header), None)
        idx_reconciliacao[nome_logico] = header.index(coluna_encontrada) if coluna_encontrada else None

    linhas = []
    for numero_linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[idx["colaborador"]] is None:
            continue  # linha em branco no fim da planilha

        data_valor = row[idx["data"]]
        if isinstance(data_valor, datetime.datetime):
            data_valor = data_valor.date().isoformat()
        elif data_valor is not None:
            data_valor = str(data_valor)[:10]

        linhas.append({
            "linha_excel": numero_linha,
            "origem": "excel",
            "colaborador": row[idx["colaborador"]],
            "data": data_valor,
            "projeto": row[idx["projeto"]],
            "idsap": str(row[idx["idsap"]]) if row[idx["idsap"]] is not None else None,
            "tarefa": row[idx["tarefa"]],
            "descricao": row[idx["descricao"]],
            "horas": float(row[idx["horas"]]) if row[idx["horas"]] is not None else None,
            "intervalo": row[idx_intervalo] if idx_intervalo is not None else None,
            "start_time": row[idx["StartTime"]],
            "end_time": row[idx["EndTime"]],
            "break_time": row[idx["Break"]],
            "employee_id_sap": int(row[idx["employeeId_sap"]]) if row[idx["employeeId_sap"]] is not None else None,
            "projeto_sap": (str(row[idx_projeto_sap]).strip() if idx_projeto_sap is not None and row[idx_projeto_sap] else None),
            **{
                nome_logico: (row[posicao] if posicao is not None else None)
                for nome_logico, posicao in idx_reconciliacao.items()
            },
        })
    return linhas
